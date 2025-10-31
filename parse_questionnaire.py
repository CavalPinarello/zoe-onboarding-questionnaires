#!/usr/bin/env python3
"""
ZOE Adaptive Onboarding - Excel Parser
Extracts questions from Sleep_Longevity_ADAPTIVE_Complete_v4.xlsx
and structures into JSON format for Supabase and visualization.
"""

import openpyxl
import json
import re
from typing import Dict, List, Any
from pathlib import Path

class QuestionnaireParser:
    def __init__(self, excel_path: str):
        self.wb = openpyxl.load_workbook(excel_path)
        self.questions = []
        self.conditional_rules = []
        self.modules = {}
        
    def parse_core_assessment(self):
        """Parse CORE Assessment sheet"""
        ws = self.wb['CORE Assessment']
        current_section = None
        question_counter = 0
        
        for row in ws.iter_rows(min_row=1, values_only=True):
            # Skip empty rows
            if not any(cell for cell in row):
                continue
                
            # Detect section headers
            if row[0] and isinstance(row[0], str) and row[0].isupper() and not row[1]:
                current_section = row[0]
                continue
            
            # Check for gateway triggers
            if row[0] and isinstance(row[0], str) and '🟠 GATEWAY' in row[0]:
                current_section = row[0].replace('🟠 GATEWAY: ', '')
                continue
            
            # Parse questions
            if row[0] and row[1]:
                try:
                    q_num = int(str(row[0]).strip())
                    question_text = row[1]
                    q_type = row[2] if len(row) > 2 else 'CORE'
                    
                    question = {
                        'id': f'CORE_{q_num}',
                        'number': q_num,
                        'text': question_text,
                        'type': q_type,
                        'section': current_section,
                        'module': 'CORE',
                        'answer_type': self._detect_answer_type(question_text),
                        'options': self._extract_options(question_text),
                        'triggers_expansion': False
                    }
                    
                    self.questions.append(question)
                    question_counter += 1
                    
                except (ValueError, AttributeError):
                    pass
            
            # Parse trigger rules
            if row[0] and isinstance(row[0], str) and '→ IF' in row[0]:
                self._parse_trigger_rule(row[0], question_counter)
        
        return self.questions
    
    def parse_expansion_module(self, sheet_name: str):
        """Parse an expansion module sheet"""
        ws = self.wb[sheet_name]
        module_name = sheet_name.replace('EXPANSION - ', '')
        module_questions = []
        
        # Get module metadata from first rows
        module_description = None
        trigger_condition = None
        
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
            if i == 0 and row[0]:
                module_description = row[0]
            if i == 1 and row[0] and 'TRIGGER' in str(row[0]).upper():
                trigger_condition = row[0]
        
        # Parse questions
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not any(cell for cell in row):
                continue
            
            if row[0] and row[1]:
                try:
                    q_num = int(str(row[0]).strip())
                    question_text = row[1]
                    q_type = row[2] if len(row) > 2 else 'EXPANSION'
                    
                    question = {
                        'id': f'{module_name.upper().replace(" ", "_")}_{q_num}',
                        'number': q_num,
                        'text': question_text,
                        'type': q_type,
                        'module': module_name,
                        'answer_type': self._detect_answer_type(question_text),
                        'options': self._extract_options(question_text),
                        'triggers_expansion': False
                    }
                    
                    module_questions.append(question)
                    
                except (ValueError, AttributeError):
                    pass
        
        if module_questions:
            self.modules[module_name] = {
                'name': module_name,
                'description': module_description,
                'trigger_condition': trigger_condition,
                'question_count': len(module_questions),
                'questions': module_questions
            }
        
        return module_questions
    
    def _detect_answer_type(self, question_text: str) -> str:
        """Detect answer type from question text"""
        text_lower = question_text.lower()
        
        if '(yes/no)' in text_lower:
            return 'boolean'
        elif '(0-10)' in text_lower or 'scale' in text_lower:
            return 'scale'
        elif 'never/rarely/sometimes/often/always' in text_lower:
            return 'frequency'
        elif 'select all' in text_lower or 'check all' in text_lower:
            return 'multiple_choice'
        elif 'inches' in text_lower or 'hours' in text_lower or 'weight' in text_lower:
            return 'numeric'
        elif 'email' in text_lower:
            return 'email'
        elif 'date of birth' in text_lower or 'date' in text_lower:
            return 'date'
        elif 'name' in text_lower:
            return 'text'
        else:
            return 'single_choice'
    
    def _extract_options(self, question_text: str) -> List[str]:
        """Extract answer options from question text"""
        # Look for options in parentheses
        matches = re.findall(r'\((.*?)\)', question_text)
        
        if matches:
            options_str = matches[-1]  # Get last parenthesis content
            
            # Split by common delimiters
            if '/' in options_str:
                return [opt.strip() for opt in options_str.split('/')]
            elif ',' in options_str:
                return [opt.strip() for opt in options_str.split(',')]
            elif options_str.startswith('0-10') or options_str.startswith('1-10'):
                return [str(i) for i in range(0, 11)] if '0-10' in options_str else [str(i) for i in range(1, 11)]
        
        return []
    
    def _parse_trigger_rule(self, rule_text: str, last_question_num: int):
        """Parse trigger rule from text"""
        # Extract condition and modules
        # Example: "→ IF YES: Expand to ISI (7 questions) + DBAS-16 (16 questions)"
        
        if '→ IF' in rule_text:
            parts = rule_text.split(':', 1)
            if len(parts) == 2:
                condition = parts[0].replace('→ IF', '').strip()
                expansions = parts[1].replace('Expand to', '').strip()
                
                # Extract module names
                modules = re.findall(r'([A-Z][A-Za-z0-9\-\s]+?)(?:\s*\(|\s*\+|$)', expansions)
                
                rule = {
                    'trigger_question_id': f'CORE_{last_question_num}',
                    'condition': condition,
                    'expanded_modules': [m.strip() for m in modules if m.strip()],
                    'rule_text': rule_text
                }
                
                self.conditional_rules.append(rule)
                
                # Mark the question as triggering expansion
                if self.questions:
                    self.questions[-1]['triggers_expansion'] = True
    
    def parse_all(self):
        """Parse all sheets"""
        print("Parsing CORE Assessment...")
        self.parse_core_assessment()
        
        print("\nParsing Expansion Modules...")
        for sheet_name in self.wb.sheetnames:
            if sheet_name.startswith('EXPANSION'):
                print(f"  - {sheet_name}")
                expansion_questions = self.parse_expansion_module(sheet_name)
                self.questions.extend(expansion_questions)
        
        return {
            'questions': self.questions,
            'conditional_rules': self.conditional_rules,
            'modules': self.modules
        }
    
    def save_json(self, output_dir: str = '.'):
        """Save parsed data to JSON files"""
        data = self.parse_all()
        
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # Save questions
        questions_file = output_path / 'questions.json'
        with open(questions_file, 'w', encoding='utf-8') as f:
            json.dump(data['questions'], f, indent=2, ensure_ascii=False)
        print(f"\n✅ Saved {len(data['questions'])} questions to {questions_file}")
        
        # Save conditional rules
        rules_file = output_path / 'conditional_rules.json'
        with open(rules_file, 'w', encoding='utf-8') as f:
            json.dump(data['conditional_rules'], f, indent=2, ensure_ascii=False)
        print(f"✅ Saved {len(data['conditional_rules'])} conditional rules to {rules_file}")
        
        # Save modules metadata
        modules_file = output_path / 'modules.json'
        with open(modules_file, 'w', encoding='utf-8') as f:
            json.dump(data['modules'], f, indent=2, ensure_ascii=False)
        print(f"✅ Saved {len(data['modules'])} modules to {modules_file}")
        
        # Create summary
        summary = {
            'total_questions': len(data['questions']),
            'core_questions': len([q for q in data['questions'] if q['module'] == 'CORE']),
            'expansion_questions': len([q for q in data['questions'] if q['module'] != 'CORE']),
            'modules': list(data['modules'].keys()),
            'conditional_rules': len(data['conditional_rules'])
        }
        
        summary_file = output_path / 'summary.json'
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2)
        print(f"✅ Saved summary to {summary_file}")
        
        return data


if __name__ == '__main__':
    excel_file = '/Users/martinkawalski/Downloads/Sleep_Longevity_ADAPTIVE_Complete_v4.xlsx'
    output_dir = '/Users/martinkawalski/ZOE/data'
    
    parser = QuestionnaireParser(excel_file)
    parser.save_json(output_dir)
